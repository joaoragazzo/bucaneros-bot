import discord
from discord import channel, colour, reaction
from discord.ext import commands
from discord.ext.commands.core import check, command
import asyncio
from asyncio.tasks import wait_for
import functools
from logging import disable, exception
import openpyxl 
from openpyxl import load_workbook
import datetime
from datetime import date, datetime
import time
from openpyxl.utils import get_column_letter
import requests
import re
import json

bot = commands.Bot(command_prefix=['ff '], description="Bot")

@bot.command(name="reportar")
async def reportar(ctx, *,nome_motivo):

    check_1 = discord.utils.get(ctx.guild.roles, id=845065650077433896)
    check_2 = discord.utils.get(ctx.guild.roles, id=853522283444437002)

    if check_1 in ctx.author.roles or check_2 in ctx.author.roles:
        pass
    else:
        sem_permissao = discord.Embed(title="⚠️ Sem permissão!", description = "Você não tem permissão para utilizar esse comando. Por favor, contate um <@&845065650077433896> ou um <&853522283444437002>.", color = discord.Colour.random())
        sem_permissao.set_footer(text="Sem permissão para utilizar esse comando!")
        message = await ctx.reply(embed = sem_permissao)
        time.sleep(10)
        await message.delete()
        return

    icon = ctx.author.avatar_url
    authorname = ctx.author.name
    nome_motivo = nome_motivo.split("/")
    nome = nome_motivo[0]
    motivo = nome_motivo[1]

    embedVar = discord.Embed(title="Reportar bolas - Bucaneiros", color = discord.Colour.random())
    embedVar.add_field(name="Atenção!", value="Esse bola já está registrado dentro na nossa lista. Deseja visualizar a ficha desse bola?")

    embedVar.set_author(name=authorname, icon_url=icon)

    reportado = discord.Embed(title="💩 Reportado com sucesso!", description=f" Foi colocado um registro contra `{nome}` por `{motivo}`", color = discord.Colour.random())

    today = date.today()
    nowtime = datetime.now()
    d1 = today.strftime("%d/%m/%Y")

    valid_reaction=["✅","❌","🆕"]
    file = "reports.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == nome:
                message = await ctx.reply(embed=embedVar)

                await message.add_reaction("✅")
                await message.add_reaction("❌")
                #await message.add_reaction("🆕")

                def check(reaction, user):
                    return user == ctx.author and str(reaction.emoji) in valid_reaction
                reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

                if str(reaction.emoji) == "✅":
                    await message.delete()
                    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
                        for cell in row:
                            if cell.value == nome:

                                motivo = ws.cell(row=cell.row, column = 4).value
                                delator = ws.cell(row=cell.row, column = 3).value
                                data = ws.cell(row=cell.row, column = 2).value
                                imagem = ws.cell(row=cell.row, column = 5).value

                                embedVar = discord.Embed(title="💩 Ficha criminal do bola", color = discord.Colour.random())
                                embedVar.add_field(name=f"Ficha abaixo", value=f"**Nickname do reportado:** `{nome}`\n**Motivo do registro: **`{motivo}`")
                                embedVar.set_footer(text=f"Denunciado por {delator} no dia {data}")
                                if imagem == None:
                                    pass
                                elif imagem != None:
                                    embedVar.set_thumbnail(url=imagem)
                                await ctx.reply(embed=embedVar)
                                return

                elif str(reaction.emoji) == "❌":
                    await message.delete()
                    return 
                
                elif str(reaction.emoji) == "🆕":
                    await message.delete()
                    
                    def check_author(msg):
                        return msg.author == ctx.author
                    
                    novo_relato = discord.Embed(title="💩 Ficha criminal do bola", description="Por favor, descreva o seu novo report.",color = discord.Colour.random())
                    novo_relato.set_footer(text=f"Adicionar novo report para `{nome}`")
                    novo_report = await ctx.reply(embed=novo_relato)

                    novo_relato = await bot.wait_for("message", timeout=60.0, check=check_author)
                    await novo_report.delete()
                    await novo_relato.delete()
                    novo_relato = novo_relato.content

                    novo_relato = discord.Embed(title="💩 Ficha cirminal do bola", description=f"Você deseja adicionar o report: `{novo_relato}` ?")
                    novo_relato.set_footer(text=f"Adicionar novo report para `{nome}`")
                    novo_relato = await ctx.reply(embed=novo_relato)

                    await novo_relato.add_reaction("✅")
                    await novo_relato.add_reaction("❌")

                    def check(reaction, user):
                        return user == ctx.author and str(reaction.emoji) in valid_reaction
                    reaction, user = await bot.wait_for('reaction_add', timeout=60.0, check=check)

                    if str(reaction.emoji) == "✅":

                        pass
                    elif str(reaction.emoji) == "❌":
                        cancelado = discord.Embed(title="❌ Novo report cancelado.", description=f"O novo report para {nome} foi cancelado. Por favor, comece outro", color = discord.Colour.random())
                        cancelado.set_footer(text="Utilize `ff reportar` novamente para começar um novo report.")
                        await ctx.reply(embed=cancelado)
                        return 

            else:
                pass
    
    ws["A{}".format(ws.max_row + 1)] = nome
    ws["B{}".format(ws.max_row)] = d1
    ws["C{}".format(ws.max_row)] = ctx.author.name
    ws["D{}".format(ws.max_row)] = motivo
    wb.save(file)
    await ctx.send(embed=reportado)

    imagem = discord.Embed(title="Informações adicionais", description="Você deseja adicionar um print do bola?", color = discord.Colour.random())
    imagem.set_footer(text=f"Adicionar informações adicionais sobre {nome}")
    imagem = await ctx.send(embed=imagem)
    await imagem.add_reaction("✅")
    await imagem.add_reaction("❌")

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)

    if str(reaction.emoji) == "✅":
        await imagem.delete()
        solicitar = discord.Embed(title="Informações adicionais", description="Por favor, me envie o print em formato de arquivo", color = discord.Colour.random())
        solicitar_ = await ctx.send(embed=solicitar)
        def check(msg):
            return msg.author == ctx.author

        imagem_url = await bot.wait_for("message", timeout = 60, check=check)
        ws["E{}".format(ws.max_row)] = imagem_url.attachments[0].url
        wb.save(file)
        await solicitar_.delete()
        
        aviso = discord.Embed(title="⚠️ Aviso!", description="Não apague o print do histórico, é por meio dele que eu terei acesso a imagem", color = discord.Colour.random())
        aviso.set_footer(text="Atenção!")
        aviso_ = await ctx.send(embed=aviso)
        time.sleep(10)
        await aviso_.delete()

    elif str(reaction.emoji) == "❌":
        await imagem.delete()
        return

@bot.command("ficha")
async def ficha(ctx, *,nome):

    file = "reports.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == nome:
                data = ws.cell(row=cell.row, column = 2).value
                delator = ws.cell(row=cell.row, column = 3).value
                motivo = ws.cell(row=cell.row, column = 4).value
                imagem = ws.cell(row=cell.row, column = 5).value

                embedVar = discord.Embed(title="💩 Ficha criminal do bola", color = discord.Colour.random())
                embedVar.add_field(name=f"Ficha abaixo", value=f"**Nickname do reportado:** `{nome}`\n**Motivo do registro:** `{motivo}`")
                embedVar.set_footer(text=f"Denunciado por {delator} no dia {data}")
                if imagem == None:
                    pass
                elif imagem != None:
                    embedVar.set_thumbnail(url=imagem)

                await ctx.reply(embed=embedVar)

@bot.command("lista")
async def lista(ctx):
    file = "reports.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    lista = ""
    n = 0
    for row in range(2, ws.max_row+1):
        for column in "A":
            n = n + 1
            cell_name="{}{}".format(column, row)
            lista = lista + f"{n}. {ws[cell_name].value}\n"
    
    lista = discord.Embed(title="💩 Lista de todos os bolas reportados", description=f"{lista}", color = discord.Colour.random())
    lista.set_footer(text="Listagem de todos os bolas reportados. Utilize ff ficha 'Nick' para obter detalhes sobre o bola.")
    await ctx.reply(embed=lista)

@bot.command("atualizar")
async def editar(ctx, *, nome):

    check_1 = discord.utils.get(ctx.guild.roles, id=845065650077433896)
    check_2 = discord.utils.get(ctx.guild.roles, id=853522283444437002)

    if check_1 in ctx.author.roles or check_2 in ctx.author.roles:
        pass
    else:
        sem_permissao = discord.Embed(title="⚠️ Sem permissão!", description = "Você não tem permissão para utilizar esse comando. Por favor, contate um <@&845065650077433896> ou um <&853522283444437002>.", color = discord.Colour.random())
        sem_permissao.set_footer(text="Sem permissão para utilizar esse comando!")
        message = await ctx.reply(embed = sem_permissao)
        time.sleep(10)
        await message.delete()
        return

    file = "reports.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == nome:
                embedVar = discord.Embed(title="💩 Atualização da ficha do bola", description=f"O que você deseja alterar no registro do {nome}?\n1️⃣ Nome do Bola\n2️⃣ Motivo do report\n3️⃣ Foto do Bola (só utilize se tiver o print já em mãos)", color = discord.Colour.random())
                edicao = await ctx.reply(embed = embedVar)
                valid_reaction = ["1️⃣","2️⃣","3️⃣"]
                await edicao.add_reaction('1️⃣')
                await edicao.add_reaction('2️⃣')
                await edicao.add_reaction('3️⃣')

                def check(reaction, user):
                    return user == ctx.author and str(reaction.emoji) in valid_reaction
                reaction, user = await bot.wait_for('reaction_add', timeout = 60, check=check)

                if str(reaction.emoji) == "1️⃣":
                    await edicao.delete()
                    solicitar = discord.Embed(title="💩 Atualização da ficha do bola", description=f"Qual o novo nick que você deseja colocar para `{nome}`?")
                    solicitar.set_footer(text=f"Atualização do nick do bola, vulgo {nome}")
                    solicitar_ = await ctx.reply(embed=solicitar)
                    def check(msg):
                        return msg.author == ctx.author
                    new_nick = await bot.wait_for("message", timeout=60, check=check)
                    ws[cell.coordinate] = new_nick.content
                    wb.save(file)
                    await solicitar_.delete()
                    await new_nick.delete()
                    mudanca = discord.Embed(title="💩 Atualização da ficha do bola", description=f"O nick foi mudado de `{nome}` para `{new_nick.content}` com sucesso!")
                    mudanca.set_footer(text="Atualização no registro realizada com sucesso!")
                    mudanca = await ctx.reply(embed=mudanca)
                    time.sleep(10)
                    await mudanca.delete()
                elif str(reaction.emoji) == "2️⃣":
                    await edicao.delete()
                    solicitar = discord.Embed(title="💩 Atualização da ficha do bola", description=f"Qual o novo report que você deseja colocar para `{nome}`?")
                    solicitar.set_footer(text=f"Atualização do report do bola, vulgo {nome}")
                    solicitar_ = await ctx.reply(embed=solicitar)
                    def check(msg):
                        return msg.author == ctx.author
                    new_report = await bot.wait_for("message", timeout=60, check=check)
                    await new_report.delete()
                    await solicitar_.delete()
                    ws[f"{get_column_letter(cell.column + 3)}{cell.row}"] = new_report.content
                    wb.save(file)
                    new_report_content = discord.Embed(title="💩 Atualização da ficha do bola", description=f"O report foi atualizado para `{new_report.content}` com sucesso!")
                    new_report_content.set_footer(text=f"Atualização do report do bola, vulgo {nome}")
                    new_report_content_ = await ctx.send(embed=new_report_content)
                    time.sleep(10)
                    await new_report_content_.delete()
                elif str(reaction.emoji) == "3️⃣":
                    await edicao.delete()
                    solicitar = discord.Embed(title="💩 Atualização da ficha do bola", description = "Por favor, envie uma foto do bola como um arquivo nesse chat.\n(⚠️ Caso queira cancelar essa ação, espere 10 segundos, caso contrário, o registro desse jogador irá quebrar! Não mande nada além de arquivos de imagem! ⚠️)")
                    solicitar.set_footer(text=f"Atualização do report do bola, vulgo {nome}")
                    await ctx.reply(embed=solicitar)
                    def check(msg):
                        return msg.author == ctx.author
                    new_photo = await bot.wait_for("message", timeout=10, check=check)
                    aviso = discord.Embed(title="⚠️ Aviso!", description="Não apague o print do histórico, é por meio dele que eu terei acesso a imagem", color = discord.Colour.random())
                    aviso.set_footer(text="Atenção!")
                    aviso = await ctx.reply(embed=aviso)
                    ws[f"{get_column_letter(cell.column + 4)}{cell.row}"] = new_photo.attachments[0].url
                    new_photo_content = discord.Embed(title="💩 Atualização da ficha do bola", description="Foi registrado uma nova foto de reconhecimento para o bola")
                    new_photo_content.set_footer(text=f"Atualização do report do bola, vulgo {nome}")
                    new_photo_content_ = await ctx.reply(embed=new_photo_content)
                    wb.save(file)
                    time.sleep(10)
                    await aviso.delete()
                    await new_photo_content_.delete()

@bot.command("ajuda")
async def ajuda(ctx):
    embedVar = discord.Embed(title="⚙️ Comando do Bot", description="Comandos do bot, suas funcionalidades e como usar!\n_Prefixo atual:_ `ff`", color = discord.Colour.random())
    embedVar.add_field(name="`reportar [Nick]/[Motivo]`", value="Adiciona um report dentro da base de dados do Bucaneros, explicitando o nome do jogador e o motivo do report. _Somente administradores._", inline=True)
    embedVar.add_field(name="`ficha [Nick]`", value="Disponibiliza a ficha do jogador em questão, caso exista.", inline = True)
    embedVar.add_field(name="`lista`", value="Mostra uma lista com todos os jogadores já reportados dentro do sistema do Bucaneros.", inline = True)
    embedVar.add_field(name="`atualizar [Nick]`", value="Atualizar informações sobre um report anteriormente já feito. _Somente administradores._", inline = True)
    embedVar.add_field(name="`status`", value="Mostra o seu status dentro do Sea of Thieves. Precisa ser cadastrado para utilizar esse comando. _O comando é pesado, portanto, pode demorar um pouco._", inline = True)
    embedVar.add_field(name="`jornada`", value="Começa seguir o seu progresso durante uma jornada no Sea of Thieves. Utilize no começo e termine o comando no final. _O comando é pesado, portanto, pode demorar um pouco._", inline=True)
    embedVar.add_field(name="`registrar`", value="Inicia o registro para ser possível utilizar os comandos `status` e `jornada`.", inline=True)
    embedVar.add_field(name="⚠️ Atenção!", value="Caso o `[Nick]` em questão não exista dentro da base de dados dos Bucaneros, o comando será ignorado, com exceção do comando `reportar`.", inline = False)
    await ctx.reply(embed = embedVar)

@bot.command("status")
async def status(ctx):

    authorid = ctx.author.id
    file = "contas.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active
    n = 0

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == authorid:
                expresssid = ws.cell(row=cell.row, column = 2).value
                ASLBSA = ws.cell(row = cell.row, column = 3).value
                ASLBSACORS = ws.cell(row = cell.row, column = 4).value
                awfs = ws.cell(row = cell.row, column = 5).value
                rat = ws.cell(row = cell.row, column = 6).value
                ApplicationGatewayAffinityCORS = ws.cell(row = cell.row, column = 7).value
                ApplicationGatewayAffinity = ws.cell(row = cell.row, column = 8).value
                n = 0
                break
            else:
                n = 1
    
    if n == 1:
        embedVar = discord.Embed(title="Usuário não cadastrado!", description="Utilize o comando `ff registrar` para começar um registro!")
        embedVar.set_footer(text="Erro por ausência de cadastro")
        await ctx.reply(embed=embedVar)
        return
    else:
        pass
                

    cookie = f"express.sid={expresssid}; ASLBSA={ASLBSA}; ASLBSACORS={ASLBSACORS}; awfs={awfs}; rat={rat}; ApplicationGatewayAffinityCORS={ApplicationGatewayAffinityCORS}; ApplicationGatewayAffinity={ApplicationGatewayAffinity}"

    cor = discord.Colour.random()
    embedVar = discord.Embed(title="Aguarde um momento...", description="<a:loading:924111683402739722> Estamos preparando tudo para você!", color = cor)
    embedVar.set_footer(text="Aguarde enquanto conectamos com o servidor do Sea of Thieves.")
    wait = await ctx.reply(embed=embedVar)

    url_balance = "https://www.seaofthieves.com/api/profilev2/balance"
    url_status = "https://www.seaofthieves.com/api/profilev2/overview?latest=3"
    headers = {
        "Cookie": f"{cookie}",
        "Sec-Ch-Ua": '" Not A;Brand";v="99", "Chromium";v="92"',
        "Sec-Ch-Ua-Mobile": "?0",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
        "Accept": "*/*",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Referer": "https://www.seaofthieves.com/profile/overview",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "en-US,en;q=0.9"
        }

    response_balance = requests.request("GET", url_balance, headers=headers)
    response_status = requests.request("GET", url_status, headers=headers)

    dados_balance = response_balance.text
    dados_status = response_status.text
    balance = dados_balance.split(",")
    status = ((dados_status.split("{"))[2]).split(",")
    
    nick_in, titulo_in, dobroes_in, gold_in, ancient_coins_in, icon_in, kraken_in, mega_in, chest_in, ship_in, miles_in = balance[0], balance[1], balance[2], balance[3], balance[4], balance[5], status[0], status[1], status[2], status[3], status[5]
    nick_tratado, titulo, dobroes, gold, ancient_coins, icon, kraken, mega, chest, ship, miles = nick_in.split(":"), titulo_in.split(":"), dobroes_in.split(":"), gold_in.split(":"), ancient_coins_in.split(":"), icon_in.split(":"), kraken_in.split(":"), mega_in.split(":"), chest_in.split(":"), ship_in.split(":"), miles_in.split(":")
    nick_final, titulo_final, icon_final = str(nick_tratado[1]).replace('"',''), str(titulo[1]).replace('"',''), str(icon[1]).replace('"','') + ":" + str(icon[2]).replace('"','')
    dobroes_num, gold_num, ancient_coins_num, kraken_num, mega_num, chest_num, ship_num, miles_num = dobroes[1], gold[1], ancient_coins[1], (str(kraken[1])).replace('"',''), (str(mega[1])).replace('"',''), str(chest[1]).replace('"',''), (str(ship[1])).replace('"',''), (str(miles[1])).replace('"','').replace("}",'')
    dobroes_final, gold_final, ancient_final, kraken_final, mega_final, chest_final, ship_final, miles_final = re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{dobroes_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{gold_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{ancient_coins_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{kraken_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{mega_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{chest_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{ship_num}"), re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{miles_num}")

    if titulo_final == "":
        titulo_final = ""
    elif titulo_final != "":
        titulo_final = f", {titulo_final}"

    embedVar = discord.Embed(title="_Perfil do Sea of Thieves_", description=f"<:gold:923891564625944597> **Ouro:** `{gold_final}`\n<:dobres:923891564839829524> **Dobrões:** `{dobroes_final}`\n<:ancient_coins:923890688184160257> **Moedas Antigas:** `{ancient_final}`\n<:krakenl:924075866890330172> **Kraken derrotados:** `{kraken_final}`\n<:megalodon:924076405149536266> **Megalodons encontrados:** `{mega_final}`\n<:chest:924077253221027851> **Baús entregados:** `{chest_final}`\n<:ship_sank:924073366489563167> **Navios afundados:** `{ship_final}`\n<:ship_miles:924075109445169162> **Milhas navegadas:** `{miles_final}`", color = cor)
    embedVar.set_footer(text=f"Mostrando o Status do jogador {nick_final}{titulo_final}")
    embedVar.set_author(name=f"{nick_final}{titulo_final}", icon_url=f"{icon_final}")
    await wait.delete()
    await ctx.reply(embed = embedVar)

@bot.command("jornada")
async def jornada(ctx):

    id = ctx.author.id

    start = datetime.now()
    start = start.strftime("%H:%M:%S")
    valid_reaction = ['✅']

    file = "contas.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active 

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == id:
                expresssid = ws.cell(row=cell.row, column = 2).value
                ASLBSA = ws.cell(row = cell.row, column = 3).value
                ASLBSACORS = ws.cell(row = cell.row, column = 4).value
                awfs = ws.cell(row = cell.row, column = 5).value
                rat = ws.cell(row = cell.row, column = 6).value
                ApplicationGatewayAffinityCORS = ws.cell(row = cell.row, column = 7).value
                ApplicationGatewayAffinity = ws.cell(row = cell.row, column = 8).value
                n = 0
                break
            else:
                n = 1
    
    if n == 1:
        embedVar = discord.Embed(title="Usuário não cadastrado!", description="Utilize o comando `ff registrar` para começar um registro!")
        embedVar.set_footer(text="Erro por ausência de cadastro")
        await ctx.reply(embed=embedVar)
        return
    else:
        pass

    cookie = f"express.sid={expresssid}; ASLBSA={ASLBSA}; ASLBSACORS={ASLBSACORS}; awfs={awfs}; rat={rat}; ApplicationGatewayAffinityCORS={ApplicationGatewayAffinityCORS}; ApplicationGatewayAffinity={ApplicationGatewayAffinity}"
    url_balance = "https://www.seaofthieves.com/api/profilev2/balance"
    url_status = "https://www.seaofthieves.com/api/profilev2/overview?latest=3"
    headers = {
        "Cookie": f"{cookie}",
        "Sec-Ch-Ua": '" Not A;Brand";v="99", "Chromium";v="92"',
        "Sec-Ch-Ua-Mobile": "?0",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
        "Accept": "*/*",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Referer": "https://www.seaofthieves.com/profile/overview",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "en-US,en;q=0.9"
        }

    response_balance = requests.request("GET", url_balance, headers=headers)
    response_status = requests.request("GET", url_status, headers=headers)

    dados_balance = response_balance.text
    dados_status = response_status.text
    balance = dados_balance.split(",")
    status = ((dados_status.split("{"))[2]).split(",")
    
    nick_in, titulo_in, dobroes_in, gold_in, ancient_coins_in, icon_in, kraken_in, mega_in, chest_in, ship_in, miles_in = balance[0], balance[1], balance[2], balance[3], balance[4], balance[5], status[0], status[1], status[2], status[3], status[5]
    nick_tratado, titulo, dobroes, gold, ancient_coins, icon, kraken, mega, chest, ship, miles = nick_in.split(":"), titulo_in.split(":"), dobroes_in.split(":"), gold_in.split(":"), ancient_coins_in.split(":"), icon_in.split(":"), kraken_in.split(":"), mega_in.split(":"), chest_in.split(":"), ship_in.split(":"), miles_in.split(":")
    nick_final, titulo_final, icon_final = str(nick_tratado[1]).replace('"',''), str(titulo[1]).replace('"',''), str(icon[1]).replace('"','') + ":" + str(icon[2]).replace('"','')
    dobroes_inicial, gold_inicial, ancient_coins_inicial, kraken_inicial, mega_inicial, chest_inicial, ship_inicial, miles_inicial = dobroes[1], gold[1], ancient_coins[1], (str(kraken[1])).replace('"',''), (str(mega[1])).replace('"',''), str(chest[1]).replace('"',''), (str(ship[1])).replace('"',''), (str(miles[1])).replace('"','').replace("}",'')
    
    if titulo_final == "":
        titulo_final = ""
    elif titulo_final != "":
        titulo_final = f", {titulo_final}"

    embedVar = discord.Embed(title="Jornada iniciada!", description="O seu progresso está sendo calculado! Ao concluir sua jornada, clique no ✅ e você receberá seu progresso!", color = discord.Colour.random())
    embedVar.set_footer(text="Seu progresso está sendo computado!")
    message = await ctx.reply(embed=embedVar)
    await message.add_reaction('✅')

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', check=check)

    if str(reaction.emoji) == '✅':
        response_balance = requests.request("GET", url_balance, headers=headers)
        response_status = requests.request("GET", url_status, headers=headers)
        dados_balance = response_balance.text
        dados_status = response_status.text
        balance = dados_balance.split(",")
        status = ((dados_status.split("{"))[2]).split(",")
    
        nick_in, titulo_in, dobroes_in, gold_in, ancient_coins_in, icon_in, kraken_in, mega_in, chest_in, ship_in, miles_in = balance[0], balance[1], balance[2], balance[3], balance[4], balance[5], status[0], status[1], status[2], status[3], status[5]
        nick_tratado, titulo, dobroes, gold, ancient_coins, icon, kraken, mega, chest, ship, miles = nick_in.split(":"), titulo_in.split(":"), dobroes_in.split(":"), gold_in.split(":"), ancient_coins_in.split(":"), icon_in.split(":"), kraken_in.split(":"), mega_in.split(":"), chest_in.split(":"), ship_in.split(":"), miles_in.split(":")
        dobroes_final, gold_final, ancient_coins_final, kraken_final, mega_final, chest_final, ship_final, miles_final = dobroes[1], gold[1], ancient_coins[1], (str(kraken[1])).replace('"',''), (str(mega[1])).replace('"',''), str(chest[1]).replace('"',''), (str(ship[1])).replace('"',''), (str(miles[1])).replace('"','').replace("}",'')

        resultado_gold = str(int(gold_final) - int(gold_inicial))
        resultado_dobroes = str(int(dobroes_final) - int(dobroes_inicial))
        resultado_ancient = str(int(ancient_coins_final) - int(ancient_coins_inicial))
        resultado_kraken = str(int(kraken_final) - int(kraken_inicial))
        resultado_mega = str(int(mega_final) - int(mega_inicial))
        resultado_chest = str(int(chest_final) - int(chest_inicial))
        resultado_ship = str(int(ship_final) - int(ship_inicial))
        resultado_miles = str(int(miles_final) - int(miles_inicial))

        ancient_final = re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{resultado_ancient}")
        gold_final = re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{resultado_gold}")
        dobroes_final = re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{resultado_dobroes}")
        miled_final = re.sub(r'(?<!^)(?=(\d{3})+$)', r'.', f"{resultado_miles}")


        end = datetime.now()
        end = end.strftime("%H:%M:%S")
        FMT = '%H:%M:%S'
        tdelta = datetime.strptime(end, FMT) - datetime.strptime(start, FMT)


        embedVar = discord.Embed(title="💎 Seu progresso da jornada!", description=f"<:gold:923891564625944597> **Ouro:** `{gold_final}`\n<:dobres:923891564839829524> **Dobrões:** `{dobroes_final}`\n<:ancient_coins:923890688184160257> **Moedas Antigas:** `{ancient_final}`\n<:krakenl:924075866890330172> **Kraken derrotados:** `{resultado_kraken}`\n<:megalodon:924076405149536266> **Megalodons encontrados:** `{resultado_mega}`\n<:chest:924077253221027851> **Baús entregados:** `{resultado_chest}`\n<:ship_sank:924073366489563167> **Navios afundados:** `{resultado_ship}`\n<:ship_miles:924075109445169162> **Milhas navegadas:** `{miled_final}`")
        embedVar.add_field(name="Informações adicionais", value=f"A sua jornada começou às `{start}` e terminou as `{end}`. No total, teve uma duração de `{tdelta}`", inline=False)
        embedVar.set_footer(text = f"Parabéns pelo seu progresso {nick_final}{titulo_final}.")
        embedVar.set_author(name=f"{nick_final}{titulo_final}", icon_url=f"{icon_final}")

        await message.delete()
        await ctx.reply(embed=embedVar)

@bot.command("registrar")
async def registrar(ctx):

    authorid = ctx.author.id
    valid_reaction = ['◀️','▶️','🆗','🔄']

    def check_reaction(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction

    def check_message(msg):
        return msg.author == ctx.author

    async def tutorial_dm():
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `express.sid`")
        await ctx.author.send(embed=embedVar)
        expresssid = await bot.wait_for("message", timeout=60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `ASLBSA`")
        await ctx.author.send(embed=embedVar)
        ASLBSA = await bot.wait_for("message", timeout=60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `ASLBSACORS`")
        await ctx.author.send(embed=embedVar)
        ASLBSACORS = await bot.wait_for("message", timeout=60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `awfs`")
        await ctx.author.send(embed=embedVar)
        awfs = await bot.wait_for("message", timeout=60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `rat`")
        await ctx.author.send(embed=embedVar)
        rat = await bot.wait_for("message", timeout=60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `ApplicationGatewayAffinityCORS`")
        await ctx.author.send(embed=embedVar)
        ApplicationGatewayAffinityCORS = await bot.wait_for("message", timeout = 60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Por favor, me envie o cookie `ApplicationGatewayAffinity`")
        await ctx.author.send(embed=embedVar)
        ApplicationGatewayAffinity = await bot.wait_for("message", timeout = 60, check=check_message)
        embedVar = discord.Embed(title="⚙️ Vincular contas Sea of Thieves | Bucaneros Crew", description="Sua conta será vinculada em breve, aguarde!")
        await ctx.author.send(embed=embedVar)
        return expresssid, ASLBSA, ASLBSACORS, awfs, rat, ApplicationGatewayAffinityCORS, ApplicationGatewayAffinity

    file = "contas.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    for row in wsr.iter_rows(wsr.min_row, wsr.max_row):
        for cell in row:
            if cell.value == authorid:

                embedVar = discord.Embed(title="⚙️ Usuário já cadastrado!", description="O seu ID já está cadastrado em nosso registro.")
                embedVar.set_footer(text="Você pode alterar seus Cookies clicando no Emoji abaixo, caso contrário, só ignore.")
                alterar = await ctx.reply(embed=embedVar)

                await alterar.add_reaction('🔄')
                reaction, user = await bot.wait_for('reaction_add', timeout=60, check=check_reaction)

                if str(reaction.emoji) == "🔄":
                    expresssid, ASLBSA, ASLBSACORS, awfs, rat, ApplicationGatewayAffinityCORS, ApplicationGatewayAffinity = await tutorial_dm()
                    ws[f"{get_column_letter(cell.column + 0)}{cell.row}"] = authorid
                    ws[f"{get_column_letter(cell.column + 1)}{cell.row}"] = expresssid.content
                    ws[f"{get_column_letter(cell.column + 2)}{cell.row}"] = ASLBSA.content
                    ws[f"{get_column_letter(cell.column + 3)}{cell.row}"] = ASLBSACORS.content
                    ws[f"{get_column_letter(cell.column + 4)}{cell.row}"] = awfs.content
                    ws[f"{get_column_letter(cell.column + 5)}{cell.row}"] = rat.content
                    ws[f"{get_column_letter(cell.column + 6)}{cell.row}"] = ApplicationGatewayAffinityCORS.content
                    ws[f"{get_column_letter(cell.column + 7)}{cell.row}"] = ApplicationGatewayAffinity.content
                    wb.save(file)
                    return                 
            else:
                pass

    cur_page = 1

    file_2 = discord.File("pagina_2.png")
    file_3 = discord.File("pagina_3.png")
    file_4 = discord.File("pagina_4.png")
    file_5 = discord.File("pagina_5.png")
    file_6 = discord.File("pagina_6.png")

    pagina_1 = discord.Embed(title="💬 Como cadastrar sua conta do Sea of Thieves?", description = "Para cadastrar a sua conta no nosso servidor, para poder utilizar o `status` e `jornada`, realize os seguintes passos!", color=discord.Colour.random())
    pagina_1.add_field(name="Atenção", value="Para conseguir vincular sua conta do Sea of Thieves com o nosso bot, siga o passo a passo cuidadosamente.", inline=True)
    pagina_1.set_footer(text="Página 1/6")

    pagina_2 = discord.Embed(title="1️⃣ Passo", description="Entre no site oficial do Sea of Thieves, já com a sua conta logada.\nEntre nesse site ⇨ https://www.seaofthieves.com/profile/overview", color = discord.Colour.random())
    pagina_2.set_image(url="attachment://pagina_2.png")
    pagina_2.set_footer(text="Página 2/6")

    pagina_3 = discord.Embed(title="2️⃣ Passo", description="Abra o 'Inspecionar Elemento' do seu navegador.\nClique com o botão direito em qualquer área do navegador e clique em 'Inspecionar Elemento'.")
    pagina_3.set_image(url="attachment://pagina_3.png")
    pagina_3.set_footer(text="Página 3/6")

    pagina_4 = discord.Embed(title="3️⃣ Passo", description="Clique em 'Aplicativo' na aba superior do menu.\nLogo depois, clique em 'Cookies' e no dominio do Sea of Thieves")
    pagina_4.set_image(url="attachment://pagina_4.png")
    pagina_4.set_footer(text="Página 4/6")

    pagina_5 = discord.Embed(title="4️⃣ Passo", description="Procure pelos valores destacados na imagem abaixo ⇨\n(`express.sid`, `ASLBSA`, `ASLBSACORS`, `awfs`, `rat`, `ApplicationGatewayAffinityCORS`, `ApplicationGatewayAffinity`)")
    pagina_5.set_image(url="attachment://pagina_5.png")
    pagina_5.set_footer(text="Página 5/6")

    pagina_6 = discord.Embed(title="5️⃣ Passo", description="Clique nos `cookies` em questão e copie no formulário que será enviado no seu privado.")
    pagina_6.set_image(url="attachment://pagina_6.png")
    pagina_6.set_footer(text="Página 6/6")

    message = await ctx.reply(embed=pagina_1)
    await message.add_reaction("▶️")

    while True:
        try:
            reaction, user = await bot.wait_for('reaction_add', timeout = 90.0, check=check_reaction)
            if str(reaction.emoji) == "▶️" and cur_page != 6:
                cur_page += 1
                if cur_page == 2:
                    await message.delete()
                    message = await ctx.reply(embed = pagina_2, file = file_2)
                    await message.add_reaction("▶️")
                if cur_page == 3:
                    await message.delete()
                    message = await ctx.reply(embed = pagina_3, file = file_3)
                    await message.add_reaction("▶️")
                if cur_page == 4:
                    await message.delete()
                    message = await ctx.reply(embed = pagina_4, file = file_4)
                    await message.add_reaction("▶️")
                if cur_page == 5:
                    await message.delete()
                    message = await ctx.reply(embed = pagina_5, file = file_5)
                    await message.add_reaction("▶️")
                if cur_page == 6:
                    await message.delete()
                    message = await ctx.reply(embed = pagina_6, file = file_6)
                    await message.add_reaction("🆗")
            elif str(reaction.emoji) == "🆗":  
                await message.delete()             
                break
        except asyncio.TimeoutError: 
            await message.delete()
            return

    expresssid, ASLBSA, ASLBSACORS, awfs, rat, ApplicationGatewayAffinityCORS, ApplicationGatewayAffinity = await tutorial_dm()
    ws[f"{get_column_letter(ws.min_column + 0)}{ws.max_row + 1}"] = authorid
    ws[f"{get_column_letter(ws.min_column + 1)}{ws.max_row}"] = expresssid.content
    ws[f"{get_column_letter(ws.min_column + 2)}{ws.max_row}"] = ASLBSA.content
    ws[f"{get_column_letter(ws.min_column + 3)}{ws.max_row}"] = ASLBSACORS.content
    ws[f"{get_column_letter(ws.min_column + 4)}{ws.max_row}"] = awfs.content
    ws[f"{get_column_letter(ws.min_column + 5)}{ws.max_row}"] = rat.content
    ws[f"{get_column_letter(ws.min_column + 6)}{ws.max_row}"] = ApplicationGatewayAffinityCORS.content
    ws[f"{get_column_letter(ws.min_column + 7)}{ws.max_row}"] = ApplicationGatewayAffinity.content
    wb.save(file)
    return

@bot.event
async def on_command_error(ctx, error):
    embedVar = discord.Embed(title="⚠️ Error", description="Comando inexistente ou utilizado da maneira incorreta!", color = discord.Colour.random())
    embedVar.set_footer(text="Utilize 'ff ajuda' para obter uma lista a respeito dos comandos!\nCódigo do erro: {}".format(error))
    message = await ctx.reply(embed=embedVar)
    await ctx.message.delete()
    time.sleep(10)
    await message.delete()

@bot.event 
async def on_ready():
    activity = discord.Game(name="e afundando bolas!", type=3)
    await bot.change_presence(status=discord.Status.online, activity=activity)
    print("<=================================================>")
    print("Bot online com sucesso:\n{0.user.name}\n{0.user.id}".format(bot))
    print("<=================================================>")

bot.run()